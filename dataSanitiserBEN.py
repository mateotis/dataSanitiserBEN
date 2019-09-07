from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(filename = 'Spring2019_BEN Participants Registration_BEN.xlsx')

sheet = wb['RAW responses - DO NOT EDIT']

numCols = sheet.max_column
numRows = sheet.max_row

# Gives names proper capitalisation
for i in range(2,numRows):
	cell = "D"+str(i)
	sheet[cell] = sheet[cell].value.lower().title() 

# Filters word content from grade; leaves only number
for i in range(2,numRows):
	cell = "E"+str(i)
	tmp = list(str(sheet[cell].value))

	for j in tmp[:]:
		if j.isnumeric() == False:
			tmp.remove(j)

	tmp2 = "".join(tmp)
	if tmp2 != "": # Sanity check
		tmp2 = int(tmp2) # Ensures appropriate formatting in Excel
	sheet[cell] = tmp2

wb.save(filename = "EDITED_Spring2019_BEN Participants Registration_BEN.xlsx")


# Load the edited file for duplicate check
wb = load_workbook(filename = 'EDITED_Spring2019_BEN Participants Registration_BEN.xlsx')

sheet = wb['RAW responses - DO NOT EDIT']

numCols = sheet.max_column
numRows = sheet.max_row

# Remove duplicate entries (checks name and grade matches)
checkList = []

for i in range(2,numRows):
	cell1 = "D"+str(i) # Name
	cell2 = "E"+str(i) # Grade
	listEntry = [sheet[cell1].value, sheet[cell2].value] # Pair of name and grade, for greater accuracy

	if listEntry not in checkList:
		checkList.append(listEntry)
	else:
		sheet.delete_rows(i, 1) # If the name is already present once, delete all subsequent appearances

checkList = []

wb.save(filename = "EDITED_Spring2019_BEN Participants Registration_BEN.xlsx")
